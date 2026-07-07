#!/usr/bin/env python3
"""生成批量状态更新 Excel 模板"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()

# ===== Sheet 1: 批量更新模板 =====
ws = wb.active
ws.title = "批量更新模板"

# 表头定义
headers = [
    ("企业名称", 30, "必填，与系统名单匹配"),
    ("统一社会信用代码", 24, "必填，作为匹配主键"),
    ("营销阶段", 22, "选填，见数据字典"),
    ("营销备注", 20, "选填，如无意向原因"),
    ("业务阶段", 18, "选填，见数据字典"),
    ("授信金额(万)", 16, "选填"),
    ("放款金额(万)", 16, "选填"),
    ("放款日期", 14, "选填，格式YYYY-MM-DD"),
    ("放款期限", 14, "选填，如12个月"),
    ("放款利率", 14, "选填，如4.85%"),
    ("产品名称", 16, "选填，如科技贷"),
    ("备注", 20, "选填"),
]

# 样式
header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="1890FF", end_color="1890FF", fill_type="solid")
required_fill = PatternFill(start_color="FFF2E8", end_color="FFF2E8", fill_type="solid")
normal_fill = PatternFill(start_color="F6FFED", end_color="F6FFED", fill_type="solid")
thin_border = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

# 写表头（第1行）
for col_idx, (name, width, tip) in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# 写说明行（第2行）
tip_font = Font(name="微软雅黑", size=9, color="8C8C8C", italic=True)
for col_idx, (_, _, tip) in enumerate(headers, 1):
    cell = ws.cell(row=2, column=col_idx, value=tip)
    cell.font = tip_font
    cell.fill = required_fill if "必填" in tip else normal_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

# 示例数据（第3-5行）
samples = [
    ["苏州智云科技有限公司", "91320500MA1EXAMPLE", "已触达-有意向", "计划下周面谈", "已申请", 500, "", "", "", "", "科技贷", ""],
    ["南京创新智造集团有限公司", "91320100MA2EXAMPLE", "已触达-无意向", "已在他行申请", "", "", "", "", "", "", "", "建行已批500万"],
    ["南通精密机械制造有限公司", "91320600MA3EXAMPLE", "", "", "已放款", 400, 100, "2026-05-01", "12个月", "4.85%", "科技贷", "第三笔放款"],
]

data_font = Font(name="微软雅黑", size=10)
for row_idx, row_data in enumerate(samples, 3):
    for col_idx, val in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = data_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center")

# 数据验证：营销阶段下拉
reach_options = '"待触达,已触达-有意向,已触达-无意向,未触达-空号/停机,未触达-拒接,未触达-非目标企业,未触达-无人接听,未触达-其他"'
dv_reach = DataValidation(type="list", formula1=reach_options, allow_blank=True)
dv_reach.error = "请从列表中选择营销阶段"
dv_reach.errorTitle = "输入错误"
dv_reach.prompt = "请选择营销阶段状态"
dv_reach.promptTitle = "营销阶段"
ws.add_data_validation(dv_reach)
dv_reach.add(f"C3:C5002")

# 数据验证：业务阶段下拉
biz_options = '"已申请,审批中,已授信,已放款,已结清,已拒绝"'
dv_biz = DataValidation(type="list", formula1=biz_options, allow_blank=True)
dv_biz.error = "请从列表中选择业务阶段"
ws.add_data_validation(dv_biz)
dv_biz.add(f"E3:E5002")

# 冻结首行
ws.freeze_panes = "A3"
ws.sheet_properties.tabColor = "1890FF"

# ===== Sheet 2: 数据字典 =====
ws2 = wb.create_sheet("数据字典")
ws2.sheet_properties.tabColor = "52C41A"

dict_headers = [("类别", 14), ("状态编码", 22), ("状态名称", 22), ("说明", 40)]
for col_idx, (name, width) in enumerate(dict_headers, 1):
    cell = ws2.cell(row=1, column=col_idx, value=name)
    cell.font = header_font
    cell.fill = PatternFill(start_color="52C41A", end_color="52C41A", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws2.column_dimensions[get_column_letter(col_idx)].width = width

reach_dict = [
    ("营销阶段", "PENDING", "待触达", "初始状态，尚未联系"),
    ("营销阶段", "REACHED_INTERESTED", "已触达-有意向", "联系成功且有融资意向"),
    ("营销阶段", "REACHED_NO_INTEREST", "已触达-无意向", "联系成功但暂无需求（营销备注填原因）"),
    ("营销阶段", "UNREACHED_EMPTY", "未触达-空号/停机", "号码无效"),
    ("营销阶段", "UNREACHED_REJECTED", "未触达-拒接", "对方拒接"),
    ("营销阶段", "UNREACHED_WRONG", "未触达-非目标企业", "号码不属于该企业"),
    ("营销阶段", "UNREACHED_NO_ANSWER", "未触达-无人接听", "多次联系不上"),
    ("营销阶段", "UNREACHED_OTHER", "未触达-其他", "其他未触达原因"),
    ("业务阶段", "APPLIED", "已申请", "客户已提交贷款申请"),
    ("业务阶段", "REVIEWING", "审批中", "银行内部审批阶段"),
    ("业务阶段", "APPROVED", "已授信", "通过审批，已出授信额度"),
    ("业务阶段", "DISBURSED", "已放款", "已放款，循环额度可多次"),
    ("业务阶段", "SETTLED", "已结清", "所有借据已结清"),
    ("业务阶段", "REJECTED_BIZ", "已拒绝", "银行拒绝授信"),
]

for row_idx, (cat, code, name, desc) in enumerate(reach_dict, 2):
    ws2.cell(row=row_idx, column=1, value=cat).font = data_font
    ws2.cell(row=row_idx, column=2, value=code).font = Font(name="Consolas", size=10)
    ws2.cell(row=row_idx, column=3, value=name).font = data_font
    ws2.cell(row=row_idx, column=4, value=desc).font = data_font

# ===== Sheet 3: 填写说明 =====
ws3 = wb.create_sheet("填写说明")
ws3.sheet_properties.tabColor = "FA8C16"
ws3.column_dimensions["A"].width = 80

instructions = [
    "【批量更新状态 - 填写说明】",
    "",
    "1. 匹配规则：系统优先以「统一社会信用代码」匹配，匹配不到时降级为「企业名称」模糊匹配。",
    "2. 必填字段：企业名称、统一社会信用代码（二者至少填一个，建议都填）。",
    "3. 状态字段：营销阶段、业务阶段均为选填，不填则不更新该维度状态。",
    "4. 金额字段：授信金额、放款金额等均为选填，不填则不更新。",
    "5. 多笔放款：同一企业若有多笔放款，可填写多行（企业名称+信用代码相同），每行填一笔放款信息。",
    "6. 单次上传上限：5000 条数据，文件格式支持 .xlsx / .csv。",
    "7. 状态值请严格使用模板中的下拉选项，自定义文本将导致解析失败。",
    "",
    "【常见问题】",
    "Q: 只想更新营销阶段，不更新业务阶段怎么办？",
    "A: 业务阶段列留空即可，系统不会覆盖已有业务阶段。",
    "",
    "Q: 一个客户有3笔放款怎么填？",
    "A: 填3行，企业名称和信用代码相同，每行的放款金额/日期/期限分别填写。",
]

title_font = Font(name="微软雅黑", bold=True, size=14, color="FA8C16")
normal_font = Font(name="微软雅黑", size=11)
for row_idx, text in enumerate(instructions, 1):
    cell = ws3.cell(row=row_idx, column=1, value=text)
    cell.font = title_font if row_idx == 1 else normal_font

output_path = "/Users/becool/Documents/联合征信/智能营销平台/营销平台Antig/bank-workbench/templates/批量更新状态_模板.xlsx"
import os
os.makedirs(os.path.dirname(output_path), exist_ok=True)
wb.save(output_path)
print(f"✅ 模板已生成: {output_path}")
